"""Report export functionality."""

import json
import logging
import csv
import io
from datetime import datetime
from typing import Dict, List, Optional, Any
from enum import Enum

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Export formats."""
    JSON = "json"
    CSV = "csv"
    HTML = "html"
    EXCEL = "excel"
    PDF = "pdf"


class ExportManager:
    """Report export management."""
    
    def __init__(self):
        """Initialize export manager."""
        self.export_history: List[Dict] = []
        self.supported_formats = [f.value for f in ExportFormat]
    
    def export_to_json(self, report_data: Dict) -> str:
        """
        Export report to JSON.
        
        Args:
            report_data: Report data
            
        Returns:
            JSON string
        """
        return json.dumps(report_data, indent=2)
    
    def export_to_csv(self, report_data: Dict) -> str:
        """
        Export report to CSV.
        
        Args:
            report_data: Report data
            
        Returns:
            CSV string
        """
        output = io.StringIO()
        
        sections = report_data.get('sections', [])
        if not sections:
            return ""
        
        # Get first section with data
        data_rows = []
        for section in sections:
            if section.get('type') == 'table' and 'rows' in section.get('content', {}):
                data_rows = section['content']['rows']
                break
        
        if not data_rows:
            return ""
        
        # Get headers from first row
        headers = list(data_rows[0].keys()) if data_rows else []
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data_rows)
        
        return output.getvalue()
    
    def export_to_html(self, report_data: Dict) -> str:
        """
        Export report to HTML.
        
        Args:
            report_data: Report data
            
        Returns:
            HTML string
        """
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{report_data.get('title', 'Report')}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #3b82f6;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
        }}
        .summary {{
            background: #f0f7ff;
            padding: 15px;
            border-left: 4px solid #3b82f6;
            margin: 20px 0;
            border-radius: 4px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th {{
            background: #3b82f6;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background: #f9f9f9;
        }}
        tr:hover {{
            background: #f0f0f0;
        }}
        .chart {{
            margin: 30px 0;
            padding: 20px;
            background: #fafafa;
            border-radius: 4px;
        }}
        .metadata {{
            color: #999;
            font-size: 12px;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{report_data.get('title', 'Report')}</h1>
        
        <div class="summary">
            <h3>Summary</h3>
            <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p><strong>Period:</strong> {report_data.get('period', 'N/A')}</p>
        </div>
"""
        
        # Add sections
        for section in report_data.get('sections', []):
            html += f"<h2>{section.get('name', 'Section')}</h2>\n"
            
            content = section.get('content', {})
            if isinstance(content, dict):
                html += "<table>\n"
                html += "<tr>"
                for key in content.keys():
                    html += f"<th>{key}</th>"
                html += "</tr>\n"
                html += "<tr>"
                for value in content.values():
                    html += f"<td>{value}</td>"
                html += "</tr>\n"
                html += "</table>\n"
        
        # Add charts as placeholders
        for chart in report_data.get('charts', []):
            html += f'<div class="chart"><p>Chart: {chart.get("title", "Chart")}</p></div>\n'
        
        html += '<div class="metadata"><p>This report was auto-generated and may contain confidential information.</p></div>'
        html += """
    </div>
</body>
</html>
"""
        return html
    
    def export_to_excel(self, report_data: Dict) -> bytes:
        """
        Export report to Excel.
        
        Note: Requires openpyxl to be installed.
        
        Args:
            report_data: Report data
            
        Returns:
            Excel file bytes
        """
        # For this implementation, we return CSV as fallback
        # In production, use openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Report"
            
            # Add title
            worksheet['A1'] = report_data.get('title', 'Report')
            worksheet['A1'].font = Font(bold=True, size=14)
            
            # Add data
            row = 3
            for section in report_data.get('sections', []):
                if section.get('type') == 'table':
                    worksheet[f'A{row}'] = section.get('name', '')
                    worksheet[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    content = section.get('content', {})
                    if isinstance(content, dict) and 'rows' in content:
                        rows = content['rows']
                        if rows:
                            # Headers
                            headers = list(rows[0].keys())
                            for col, header in enumerate(headers, 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.value = header
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            row += 1
                            
                            # Data
                            for record in rows:
                                for col, header in enumerate(headers, 1):
                                    cell = worksheet.cell(row=row, column=col)
                                    cell.value = record.get(header, '')
                                row += 1
                    row += 1
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            return output.getvalue()
        
        except ImportError:
            logger.warning("openpyxl not installed, using CSV fallback")
            return self.export_to_csv(report_data).encode('utf-8')
    
    def export_to_pdf(self, report_data: Dict) -> bytes:
        """
        Export report to PDF.
        
        Note: Requires pdfkit and wkhtmltopdf to be installed.
        
        Args:
            report_data: Report data
            
        Returns:
            PDF file bytes
        """
        # For this implementation, convert HTML to PDF
        # In production, use pdfkit with wkhtmltopdf
        try:
            import pdfkit
            
            html_content = self.export_to_html(report_data)
            pdf_bytes = pdfkit.from_string(html_content, False)
            return pdf_bytes
        
        except ImportError:
            logger.warning("pdfkit not installed, using HTML fallback")
            return self.export_to_html(report_data).encode('utf-8')
    
    def export_report(self, report_data: Dict, format: str) -> Dict:
        """
        Export report in specified format.
        
        Args:
            report_data: Report data
            format: Export format
            
        Returns:
            Export result with file content
        """
        if format == ExportFormat.JSON.value:
            content = self.export_to_json(report_data)
            filename = f"{report_data.get('id', 'report')}.json"
        elif format == ExportFormat.CSV.value:
            content = self.export_to_csv(report_data)
            filename = f"{report_data.get('id', 'report')}.csv"
        elif format == ExportFormat.HTML.value:
            content = self.export_to_html(report_data)
            filename = f"{report_data.get('id', 'report')}.html"
        elif format == ExportFormat.EXCEL.value:
            content = self.export_to_excel(report_data)
            filename = f"{report_data.get('id', 'report')}.xlsx"
        elif format == ExportFormat.PDF.value:
            content = self.export_to_pdf(report_data)
            filename = f"{report_data.get('id', 'report')}.pdf"
        else:
            return {'status': 'error', 'message': f'Unsupported format: {format}'}
        
        # Log export
        self.export_history.append({
            'report_id': report_data.get('id'),
            'format': format,
            'timestamp': datetime.now().isoformat(),
            'filename': filename
        })
        
        logger.info(f"Report exported: {filename}")
        
        return {
            'status': 'success',
            'filename': filename,
            'format': format,
            'size': len(content) if isinstance(content, bytes) else len(content.encode())
        }
    
    def get_stats(self) -> Dict:
        """Get export statistics."""
        return {
            'total_exports': len(self.export_history),
            'supported_formats': self.supported_formats,
            'format_breakdown': {}
        }
